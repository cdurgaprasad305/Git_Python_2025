


from openpyxl import Workbook

# Writing values in excel sheet

wb = Workbook()
ws = wb.active
ws['A1'] = "Country"
ws['b1'] = "Capital"
ws['C1'] = "Language"

ws.cell(2,2,"India")
# It return value in B2

print (ws.cell(1,1).value)
# First row first column value
print ("Total no.of Rws " + str(ws.max_row))
print("Total No.of Columns " + str(ws.max_column))
wb.save("Third.xls")

print (dir(wb))
print ("===========================================")
print (dir(ws))


# op:
# Country
# Total no.of Rows 2
# Total No.of Columns 3

